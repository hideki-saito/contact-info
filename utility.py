import glob
import ntpath
import json
import ast

from openpyxl import load_workbook
import pandas as pd
from pandas import ExcelWriter

from google import search
from config import *


def get_status():
    try:
        with open(os.path.join(rootpath, 'status', "status.json")) as f:
            data = json.load(f)
            return data['sheet'], data['index']
    except:
        return None, 0

def write_status(sheetName, index):
    with open(os.path.join(rootpath, 'status', "status.json"), 'w') as f:
        data = {'sheet': sheetName, 'index': index}
        json.dump(data, f)


# print (get_status())
# write_status("webinar", 10)
# print (get_status())


def add_dfToexcel(df, sheetname, columns):
    output_file = output

    try:
        last_row = len(pd.read_excel(output_file, sheetname))
        book = load_workbook(output_file)
        writer = ExcelWriter(output_file, engine='openpyxl')
        writer.book = book
        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

        df.to_excel(writer, sheetname, startrow=last_row + 1, header=False, columns=columns,
                    index=False)
        # df.to_excel(writer, sheetname, startrow=last_row + 1, header=False, columns=['c', 'a'],
        #             index=False)

        writer.save()

    except:

        book = load_workbook(output_file)
        writer = ExcelWriter(output_file, engine='openpyxl')
        writer.book = book
        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
        # emtpy_df = pd.DataFrame(columns=columns)
        df.to_excel(writer, sheetname, header=True, columns=columns,
                    index=False)

        writer.save()


# a = pd.Series({'c':'b', 'd':'d'})
# df = pd.DataFrame([a], columns=list(a.keys()))
# add_dfToexcel(df, 'test', ['a', 'c'])


def add_dict(json_data, txtname):
    path = os.path.join(rootpath, 'output', txtname)
    if not os.path.exists(path):
        with open(path, 'w') as f:
            # json.dump(data, f)
            f.write(str(json_data))
    else:
        with open(path, 'a') as f:
            f.write("\n")
            # json.dump(data, f)
            f.write(str(json_data))


def unscanned_starting(txtname, emails):
    path = os.path.join(rootpath, 'output', txtname)
    if os.path.exists(path):
        with open(path) as f:
            txt_data = f.read()
        count = 0
        while count < len(emails):
            if emails[count] in txt_data:
                count += 1
            else:
                return count
    else:
        return 0


def read_dict(txtname):
    import re
    # path = os.path.join(rootpath, 'output', txtname)
    path = txtname
    data_list = []
    if os.path.exists(path):
        with open(path, 'r') as f:
            lines = list(set(re.findall("\{.*?\}", f.read().replace("\n", ""))))
            for line in lines:
                line = line.strip("\n")
                data = ast.literal_eval(line)
                data.pop('', None)
                data_list.append(data)

    return data_list


def google_search(keyword):
    for url in search(keyword, stop=20):
        print(url)


def get_keyword1(df):
    return df['First Name'] + " " + df['Last Name'] + ' at ' + df['Company / Account']


def excel_addColumns(filename, column_data):
    from openpyxl import load_workbook
    from pandas import ExcelWriter
    book = load_workbook(filename)
    writer = ExcelWriter(filename, engine='openpyxl')
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)


def find_similarity(data_list):
    pairs = []
    for item in data_list:
        if item.strip('S') in data_list and item.strip('S') != item:
            pairs.append((item, item.strip('S')))

    return pairs


def polish_data(data):
    df = pd.DataFrame(data)
    df.fillna("", inplace=True)
    new_columns = [item.strip(":") for item in list(df.columns.values)]
    df.columns = new_columns

    merge_columnPairs = find_similarity(list(df.columns.values))
    # print (merge_columnPairs)
    for item in merge_columnPairs:
        df[item[1]] = df[item[1]] + df[item[0]]

        del df[item[0]]

    return df


def makeExcel_fromTxt():
    txt_files = glob.glob(os.path.join(rootpath, 'output', '*.txt'))
    for txt_file in txt_files[:1]:
        data = read_dict(txt_file)
        extraDf = polish_data(data)
        print (list(extraDf.columns))
        # sheetname = ntpath.basename(txt_file).split("_more")[0]
        # df = pd.read_excel(source_file, sheetname)
        # new_df = pd.merge(df, extraDf, on='email')

# makeExcel_fromTxt()

# google_search("Ariel Crohn at Coverys")
# https://pipl.com/search/?q=jimmy.bourdon%40parexel.com&in=5&l=&sloc=

# make_doc("webina", dict_data)


# a = {'a':'b', 'c':'d'}
# b = {'e':'f', 'g':'h'}
# add_json(a, 'a.txt')
# add_json(b, 'a.txt')
# read_dict("a.txt")
